
## Script to Convert Excel (w Multiple sheets) into Fortigate Static Route config.
## Excel contains 3 Columns "Destination", "Nexthop" & "Interface".
## Script convert values under the columns into Fortigate static route config
##

# Import openpyxl library
import openpyxl

# Prompt the user to input the file path of the excel file
file_path = input("Enter the file path of the excel file: ")

# Load the excel workbook and get the sheet names
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet_names = wb.sheetnames

# Create an empty list to store the fortigate configurations
config_list = []

# Loop through each sheet name
for sheet_name in sheet_names:
    # Get the sheet object
    sheet = wb[sheet_name]
    # Add the sheet name as a comment to the config list
    config_list.append(f"# {sheet_name}")
    # Get the header row values
    header_row = [cell.value for cell in sheet[1]]
    # Check if the header row contains "Destination", "Nexthop" and "Interface" columns
    if "Destination" in header_row and "Nexthop" in header_row and "Interface" in header_row and "Comment" in header_row:
        # Get the column indices of "Destination", "Nexthop" and "Interface"
        dst_col = header_row.index("Destination") + 1
        nh_col = header_row.index("Nexthop") + 1
        itf_col = header_row.index("Interface") +1
        idx_col = header_row.index("Index") +1
        cm_col = header_row.index("Comment") + 1
        # Loop through each row of the sheet, starting from the second row (skipping the header row)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Get the object and IP values from the row
            dst = row[dst_col - 1]
            nh = row[nh_col - 1]
            itf = row[itf_col - 1]
            idx = row[idx_col - 1]
            cm = row[cm_col - 1]
            # Check if both the object name and IP address are not empty
            if dst is not None and nh is not None and int is not None and cm is not None:
                # Create a fortigate address object configuration string
                config_string = f"config router static\n" \
                                f"edit {idx}\n" \
                                f"set dst {dst}\n" \
                                f"set gateway {nh}\n" \
                                f"set device \"{itf}\"\n" \
                                f"set comment \"{cm}\"\n" \
                                f"next\n" \
                                f"end"
            # Add the config string to the config list
            config_list.append(config_string)
    else:
        # Print a message to indicate that the sheet does not have the required columns
        print(f"The sheet {sheet_name} does not have 'Destination', 'Nexthop' or 'Interface' columns. Skipping this sheet.")

# Join the config list with newlines and store it as a variable
config_output = "\n".join(config_list)

# Write the config output to a text file
with open("Fg_Static_Route.txt", "w") as f:
    f.write(config_output)

# Print a message to indicate the script is done
print("The script is done. Check the Fg_Static_Route.txt file for the results.")
