
## Script to Convert Excel (w Multiple sheets) into Fortigate Address Objects config.
## Excel contains 2 Columns "Object" & "IP".
## Script convert "Object" values into Address object
## And "IP" values as ip-address.
##


# Import openpyxl library
import openpyxl

# Prompt the user to input the file path of the excel file
file_path = input("Enter the file path of the excel file: ")

# Load the excel workbook and get the sheet names
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet_names = wb.sheetnames

# Create an empty list to store the fortigate address object configurations
config_list = []

# Loop through each sheet name
for sheet_name in sheet_names:
    # Get the sheet object
    sheet = wb[sheet_name]
    # Add the sheet name as a comment to the config list
    config_list.append(f"# {sheet_name}")
    # Get the header row values
    header_row = [cell.value for cell in sheet[1]]
    # Check if the header row contains "Object" and "IP" columns
    if "addrobj" in header_row and "ip" in header_row:
        # Get the column indices from header
        ao_col = header_row.index("addrobj") + 1
        ip_col = header_row.index("ip") + 1
        mk_col = header_row.index("mask") + 1
        cm_col = header_row.index("comment") + 1
        # Loop through each row of the sheet, starting from the second row (skipping the header row)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Get the object and IP values from the row
            addrobj = row[ao_col - 1]
            ip = row[ip_col - 1]
            mask = row[mk_col - 1]
            comment = row[cm_col - 1]
            # Check if both the object name and IP address are not empty
            if addrobj is not None and ip is not None:
                # Create a fortigate address object configuration string
                config_string = f"config firewall address\n" \
                                f"edit \"{addrobj}\"\n" \
                                f"set subnet {ip} {mask}\n" \
                                f"set comment \"{comment}\"\n" \
                                f"next\n" \
                                f"end"
            # Add the config string to the config list
            config_list.append(config_string)
    elif "addrrobj" in header_row and "type" in header_row:
        # Get the column indices from header
        ar_col = header_row.index("addrrobj") + 1
        ty_col = header_row.index("type") + 1
        si_col = header_row.index("start-ip") + 1
        ei_col = header_row.index("end-ip") + 1
        cm_col = header_row.index("comment") + 1
        # Loop through each row of the sheet, starting from the second row (skipping the header row)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Get the object and IP values from the row
            addrrobj = row[ar_col - 1]
            type = row[ty_col - 1]
            startip = row[si_col - 1]
            endip = row[ei_col - 1]
            comment = row[cm_col - 1]
            # Check if both the object name and IP address are not empty
            if addrrobj is not None and type is not None:
                # Create a fortigate address object configuration string
                config_string = f"config firewall address\n" \
                                f"edit \"{addrrobj}\"\n" \
                                f"set start-ip {startip}\n" \
                                f"set end-ip {endip}\n" \
                                f"set comment \"{comment}\"\n" \
                                f"next\n" \
                                f"end"
            # Add the config string to the config list
            config_list.append(config_string)
    elif "svcobj" in header_row and "protocol" in header_row:
        # Get the column indices from header
        so_col = header_row.index("svcobj") + 1
        pr_col = header_row.index("protocol") + 1
        pt_col = header_row.index("port") + 1
        cm_col = header_row.index("comment") + 1
        # Loop through each row of the sheet, starting from the second row (skipping the header row)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Get the object and IP values from the row
            svcobj = row[so_col - 1]
            protocol = row[pr_col - 1]
            port = row[pt_col - 1]
            comment = row[cm_col - 1]
            # Check if both the object name and IP address are not empty
            if svcobj is not None and protocol is not None:
                # Create a fortigate address object configuration string
                config_string = f"config firewall service custom\n" \
                                f"edit \"{svcobj}\"\n" \
                                f"set {protocol}-portrange {port}\n" \
                                f"set comment \"{comment}\"\n" \
                                f"next\n" \
                                f"end"
            # Add the config string to the config list
            config_list.append(config_string)
    else:
        # Print a message to indicate that the sheet does not have the required columns
        print(f"The sheet {sheet_name} does not have 'Objects' and 'Additional' columns. Skipping this sheet.")

# Join the config list with newlines and store it as a variable
config_output = "\n".join(config_list)

# Write the config output to a text file
with open("FG_Obj_Conf.txt", "w") as f:
    f.write(config_output)

# Print a message to indicate the script is done
print("The script is done. Check the FG_Obj_Conf.txt file for the results.")
