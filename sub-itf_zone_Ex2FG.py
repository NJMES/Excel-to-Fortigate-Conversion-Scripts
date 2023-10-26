
## Script to Convert Excel (w Multiple sheets) into Fortigate Static Route config.
## Excel will contains Columns (name, ip, subnet, interface, vlan-id, subzone, vlanforward, device-id, role)
## Script convert values under the columns into Fortigate zone with sub-interfaces.
##

# Import openpyxl library
import openpyxl

# Prompt the user to input the file path of the excel file
file_path = input("Enter the file path of the excel file: ")

# Load the excel workbook and get the sheet names
wb = openpyxl.load_workbook(file_path, data_only=True)
sheet_names = wb.sheetnames

# Create an empty list to store the fortigate configurations
config_list = []

# Loop through each sheet name
for sheet_name in sheet_names:
    # Get the sheet object
    sheet = wb[sheet_name]
    # Add the sheet name as a comment to the config list
    config_list.append(f"# {sheet_name}")
    # Get the header row values
    header_row = [cell.value for cell in sheet[1]]
    # Check if the header row contains "values" columns
    if "name" in header_row and "ip" in header_row and "interface" in header_row and "vlanid" in header_row:
        # headers (name,ip,subnet,interface,vlanid,subzone,vlanforward,device-id,role)
        nm_col = header_row.index("name") + 1
        vd_col = header_row.index("vdom") + 1
        ip_col = header_row.index("ip") + 1
        sn_col = header_row.index("subnet") +1
        itf_col = header_row.index("interface") +1
        vid_col = header_row.index("vlanid") + 1
        sz_col = header_row.index("subzone") + 1
        vf_col = header_row.index("vlanforward") + 1
        aa_col = header_row.index("allowaccess") + 1
        did_col = header_row.index("device-id") + 1
        rl_col = header_row.index("role") + 1

        # Loop through each row of the sheet, starting from the second row (skipping the header row)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Get the values from the row
            # headers (name,ip,subnet,interface,vlanid,subzone,vlanforward,device-id,role)
            nm = row[nm_col - 1]
            vd = row[vd_col -1]
            ip = row[ip_col - 1]
            sn = row[sn_col - 1]
            itf = row[itf_col - 1]
            vid = row[vid_col - 1]
            sz = row[sz_col - 1]
            vf = row[vf_col - 1]
            aa = row[aa_col -1]
            did = row[did_col - 1]
            rl = row[rl_col - 1]
            # Check if both the object name and IP address are not empty
            if nm is not None and ip is not None and vid is not None and sz is not None:
                # Create a fortigate Zone & Interface configuration string

                config_string = f"config system interface\n" \
                                f"edit V{vid}\n" \
                                f"set vdom {vd}\n" \
                                f"set vlanforward {vf}\n" \
                                f"set device-identification {did}\n" \
                                f"set allowaccess {aa}\n" \
                                f"set role {rl}\n" \
                                f"set ip {ip} {sn}\n" \
                                f"set interface {itf}\n" \
                                f"set vlanid {vid}\n" \
                                f"next\n" \
                                f"end\n" \
                                f"config system zone\n" \
                                f"edit {nm}\n" \
                                f"set interface V{vid}\n" \
                                f"next\n" \
                                f"end"
            # Add the config string to the config list
            config_list.append(config_string)
    else:
        # Print a message to indicate that the sheet does not have the required columns
        print(f"The sheet {sheet_name} does not have correct 'interface info' columns. Skipping this sheet.")

# Join the config list with newlines and store it as a variable
config_output = "\n".join(config_list)

# Write the config output to a text file
with open("FG_Itf_Zone.txt", "w") as f:
    f.write(config_output)

# Print a message to indicate the script is done
print("The script is done. Check the FG_Itf_Zone.txt file for the results.")
