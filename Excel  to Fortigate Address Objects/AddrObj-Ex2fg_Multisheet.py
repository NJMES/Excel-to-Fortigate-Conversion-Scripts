# Import openpyxl library
import openpyxl

# Prompt the user to input the file path of the excel file
file_path = input("Enter the file path of the excel file: ")

# Load the excel workbook and get the sheet names
wb = openpyxl.load_workbook(file_path)
sheet_names = wb.sheetnames

# Create an empty list to store the fortigate address object configurations
config_list = []

# Loop through each sheet name
for sheet_name in sheet_names:
    # Get the sheet object
    sheet = wb[sheet_name]
    # Add the sheet name as a comment to the config list
    config_list.append(f"# {sheet_name}")
    # Get the header row values
    header_row = [cell.value for cell in sheet[1]]
    # Check if the header row contains "Object" and "IP" columns
    if "Object" in header_row and "IP" in header_row:
        # Get the column indices of "Object" and "IP"
        object_col = header_row.index("Object") + 1
        ip_col = header_row.index("IP") + 1
        # Loop through each row of the sheet, starting from the second row (skipping the header row)
        for row in sheet.iter_rows(min_row=2):
            # Get the object and IP values from the row
            object = row[object_col - 1].value
            ip = row[ip_col - 1].value
            # Create a fortigate address object configuration string
            config_string = f"edit {object}\nset subnet {ip}\nnext"
            # Add the config string to the config list
            config_list.append(config_string)
    else:
        # Print a message to indicate that the sheet does not have the required columns
        print(f"The sheet {sheet_name} does not have 'Object' and 'IP' columns. Skipping this sheet.")

# Join the config list with newlines and store it as a variable
config_output = "\n".join(config_list)

# Write the config output to a text file
with open("config_output.txt", "w") as f:
    f.write(config_output)

# Print a message to indicate the script is done
print("The script is done. Check the config_output.txt file for the results.")
